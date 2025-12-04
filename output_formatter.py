from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

class OutputFormatter:
    """输出格式化模块，将搜索结果整理为美观易读的格式"""
    
    def __init__(self):
        pass
    
    def format_results(self, results_dict):
        """
        格式化搜索结果
        :param results_dict: 搜索结果字典，key为关键词，value为解析后的结果列表
        :return: 格式化后的字符串
        """
        if not results_dict:
            return "没有搜索结果"
        
        formatted_output = """
==============================================
自动化储能信息源搜索结果
==============================================
"""
        
        for keyword, results in results_dict.items():
            formatted_output += f"\n【搜索关键词】: {keyword}\n"
            formatted_output += f"【结果数量】: {len(results)}\n"
            formatted_output += "-" * 50
            formatted_output += "\n"
            
            for i, result in enumerate(results):
                formatted_output += f"\n{str(i+1)}. 【标题】: {result['title']}\n"
                formatted_output += f"   【部分内容】: {result['content']}\n"
                formatted_output += f"   【链接】: {result['url']}\n"
                formatted_output += "   " + "-" * 45 + "\n"
        
        formatted_output += "\n==============================================\n"
        formatted_output += "搜索完成\n"
        formatted_output += "=============================================="
        
        return formatted_output
    
    def save_to_file(self, results_dict, filename="search_results.txt"):
        """
        将搜索结果保存到文件
        :param results_dict: 搜索结果字典
        :param filename: 保存的文件名
        """
        formatted_content = self.format_results(results_dict)
        
        try:
            with open(filename, "w", encoding="utf-8") as f:
                f.write(formatted_content)
            print(f"搜索结果已保存到文件: {filename}")
        except Exception as e:
            print(f"保存搜索结果到文件时出错: {str(e)}")
    
    def save_to_excel(self, results_dict, filename="search_results.xlsx"):
        """
        将搜索结果保存到Excel文件
        :param results_dict: 搜索结果字典
        :param filename: 保存的文件名
        """
        if not results_dict:
            print("没有搜索结果可保存到Excel")
            return
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "搜索结果"
        
        # 写入表头，增加页面分析和储能分类相关列
        headers = ["搜索关键词", "序号", "标题", "部分内容", "链接地址", "跳转链接", 
                  "是否储能领域", "储能类别", "公司类型", "置信度", "分析理由", "页面标题", 
                  "页面字数", "页面分析状态", "LLM分析状态"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            worksheet[f"{col_letter}1"] = header
            # 设置表头居中对齐
            worksheet[f"{col_letter}1"].alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据行
        row_num = 2
        for keyword, results in results_dict.items():
            for idx, result in enumerate(results, 1):
                worksheet.cell(row=row_num, column=1, value=keyword)
                worksheet.cell(row=row_num, column=2, value=idx)
                worksheet.cell(row=row_num, column=3, value=result["title"])
                worksheet.cell(row=row_num, column=4, value=result["content"])
                worksheet.cell(row=row_num, column=5, value=result["url"])
                worksheet.cell(row=row_num, column=6, value=result.get("redirect_url", ""))
                
                # 添加页面分析和储能分类相关信息
                worksheet.cell(row=row_num, column=7, value="是" if result.get("is_energy_storage", False) else "否")
                worksheet.cell(row=row_num, column=8, value=result.get("storage_category", ""))
                worksheet.cell(row=row_num, column=9, value=result.get("company_type", ""))
                worksheet.cell(row=row_num, column=10, value=result.get("confidence", ""))
                worksheet.cell(row=row_num, column=11, value=result.get("analysis_reason", ""))
                
                # 添加页面元数据
                if "page_metadata" in result and isinstance(result["page_metadata"], dict):
                    worksheet.cell(row=row_num, column=12, value=result["page_metadata"].get("title", ""))
                    worksheet.cell(row=row_num, column=13, value=result["page_metadata"].get("word_count", ""))
                else:
                    worksheet.cell(row=row_num, column=12, value="")
                    worksheet.cell(row=row_num, column=13, value="")
                
                # 添加分析状态信息
                worksheet.cell(row=row_num, column=14, value="失败: " + result.get("page_error", "") if "page_error" in result else "成功")
                worksheet.cell(row=row_num, column=15, value="失败: " + result.get("llm_error", "") if "llm_error" in result else "成功" if "is_energy_storage" in result else "未分析")
                
                row_num += 1
        
        # 自动调整列宽
        for col_num in range(1, len(headers) + 1):
            max_width = 0
            column_letter = get_column_letter(col_num)
            
            # 计算该列最大宽度
            for cell in worksheet[column_letter]:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_width:
                        max_width = cell_length
            
            # 为特定列设置默认宽度
            if column_letter in ['D', 'J']:  # 部分内容和分析理由列
                default_width = 50
            elif column_letter in ['E']:  # 链接地址列
                default_width = 40
            elif column_letter in ['K']:  # 页面标题列
                default_width = 40
            else:
                default_width = 20
            
            # 设置列宽（取最大宽度+2和默认宽度的较大值，但不超过50）
            worksheet.column_dimensions[column_letter].width = min(max(max_width + 2, default_width), 50)
        
        try:
            workbook.save(filename)
            print(f"搜索结果已保存到Excel文件: {filename}")
        except Exception as e:
            print(f"保存搜索结果到Excel时出错: {str(e)}")
    
    def print_results(self, results_dict):
        """
        打印搜索结果
        :param results_dict: 搜索结果字典
        """
        formatted_content = self.format_results(results_dict)
        print(formatted_content)